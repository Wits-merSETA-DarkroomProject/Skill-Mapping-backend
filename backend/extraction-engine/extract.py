"""
File Content Extraction Script
Extracts text and metadata from various file types (PDF, DOCX, XLSX, TXT, etc.)
Also saves extracted content to files and database
"""

import os
import json
import sqlite3
from pathlib import Path
from typing import Dict, Any, Optional, Union
from datetime import datetime
import hashlib


class FileExtractor:
    """Main class for extracting content from different file types."""
    
    SUPPORTED_FORMATS = {
        '.pdf': 'PDF Document',
        '.txt': 'Text File',
        '.docx': 'Word Document',
        '.doc': 'Word Document',
        '.xlsx': 'Excel Spreadsheet',
        '.xls': 'Excel Spreadsheet',
        '.csv': 'CSV File',
        '.json': 'JSON File',
        '.html': 'HTML File',
        '.md': 'Markdown File'
    }
    
    def __init__(self, upload_dir: str = "uploads", results_dir: str = "extraction_results"):
        """Initialize the extractor with upload and results directories."""
        self.upload_dir = upload_dir
        self.results_dir = results_dir
        self._ensure_upload_dir()
        self._ensure_results_dir()
        self._init_database()
    
    def _ensure_upload_dir(self):
        """Create upload directory if it doesn't exist."""
        Path(self.upload_dir).mkdir(parents=True, exist_ok=True)
    
    def _ensure_results_dir(self):
        """Create results directory structure if it doesn't exist."""
        Path(self.results_dir).mkdir(parents=True, exist_ok=True)
        Path(f"{self.results_dir}/json").mkdir(parents=True, exist_ok=True)
        Path(f"{self.results_dir}/text").mkdir(parents=True, exist_ok=True)
        Path(f"{self.results_dir}/database").mkdir(parents=True, exist_ok=True)
    
    def _init_database(self):
        """Initialize SQLite database for storing extraction records."""
        db_path = f"{self.results_dir}/database/extractions.db"
        self.db_path = db_path
        
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # Create table for extraction records
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS extractions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id TEXT UNIQUE NOT NULL,
                    file_name TEXT NOT NULL,
                    file_type TEXT,
                    file_path TEXT,
                    file_size_bytes INTEGER,
                    content TEXT,
                    content_summary TEXT,
                    metadata TEXT,
                    extracted_at TIMESTAMP,
                    saved_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status TEXT DEFAULT 'success'
                )
            ''')
            
            # Create index on file_id for faster lookups
            cursor.execute('''
                CREATE INDEX IF NOT EXISTS idx_file_id ON extractions(file_id)
            ''')
            
            # Create index on extracted_at for date queries
            cursor.execute('''
                CREATE INDEX IF NOT EXISTS idx_extracted_at ON extractions(extracted_at)
            ''')
            
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Database initialization warning: {str(e)}")
    
    def extract(self, file_path: str) -> Dict[str, Any]:
        """
        Extract content from any supported file type.
        
        Args:
            file_path: Path to the file to extract
            
        Returns:
            Dictionary containing extracted content and metadata
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            return {"error": f"File not found: {file_path}", "success": False}
        
        file_ext = file_path.suffix.lower()
        
        if file_ext not in self.SUPPORTED_FORMATS:
            return {
                "error": f"Unsupported file type: {file_ext}",
                "supported_formats": list(self.SUPPORTED_FORMATS.keys()),
                "success": False
            }
        
        extraction_method = getattr(self, f'_extract{file_ext.replace(".", "_")}', None)
        
        if extraction_method is None:
            return {"error": f"No extractor implemented for {file_ext}", "success": False}
        
        try:
            content = extraction_method(file_path)
            metadata = self._get_file_metadata(file_path)
            
            return {
                "success": True,
                "file_name": file_path.name,
                "file_type": self.SUPPORTED_FORMATS.get(file_ext, "Unknown"),
                "file_size_bytes": file_path.stat().st_size,
                "content": content,
                "metadata": metadata,
                "extracted_at": datetime.now().isoformat()
            }
        except Exception as e:
            return {
                "error": f"Error extracting from {file_path.name}: {str(e)}",
                "success": False
            }
    
    def _extract_pdf(self, file_path: Path) -> str:
        """Extract text from PDF files."""
        try:
            import PyPDF2
            
            text_content = []
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    text = page.extract_text()
                    if text.strip():
                        text_content.append(f"--- Page {page_num} ---\n{text}")
            
            return "\n\n".join(text_content) if text_content else "No text found in PDF"
        except ImportError:
            return "ERROR: PyPDF2 not installed. Install with: pip install PyPDF2"
        except Exception as e:
            raise Exception(f"PDF extraction failed: {str(e)}")
    
    def _extract_txt(self, file_path: Path) -> str:
        """Extract text from plain text files."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            # Try with different encoding
            with open(file_path, 'r', encoding='latin-1') as file:
                return file.read()
    
    def _extract_docx(self, file_path: Path) -> str:
        """Extract text from Word (.docx) files."""
        try:
            from docx import Document
            
            doc = Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs) if paragraphs else "No content found"
        except ImportError:
            return "ERROR: python-docx not installed. Install with: pip install python-docx"
        except Exception as e:
            raise Exception(f"DOCX extraction failed: {str(e)}")
    
    def _extract_doc(self, file_path: Path) -> str:
        """Extract text from legacy Word (.doc) files."""
        try:
            import docx2txt
            return docx2txt.process(file_path)
        except ImportError:
            return "ERROR: docx2txt not installed. Install with: pip install docx2txt"
        except Exception as e:
            raise Exception(f"DOC extraction failed: {str(e)}")
    
    def _extract_xlsx(self, file_path: Path) -> Union[str, Dict]:
        """Extract data from Excel files."""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path)
            result = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(row)
                result[sheet_name] = sheet_data
            
            return json.dumps(result, indent=2, default=str)
        except ImportError:
            return "ERROR: openpyxl not installed. Install with: pip install openpyxl"
        except Exception as e:
            raise Exception(f"XLSX extraction failed: {str(e)}")
    
    def _extract_xls(self, file_path: Path) -> Union[str, Dict]:
        """Extract data from legacy Excel files."""
        try:
            import xlrd
            
            workbook = xlrd.open_workbook(file_path)
            result = {}
            
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                sheet_data = []
                for row_idx in range(sheet.nrows):
                    sheet_data.append(sheet.row_values(row_idx))
                result[sheet_name] = sheet_data
            
            return json.dumps(result, indent=2, default=str)
        except ImportError:
            return "ERROR: xlrd not installed. Install with: pip install xlrd"
        except Exception as e:
            raise Exception(f"XLS extraction failed: {str(e)}")
    
    def _extract_csv(self, file_path: Path) -> str:
        """Extract data from CSV files."""
        try:
            import csv
            
            rows = []
            with open(file_path, 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                rows = list(reader)
            
            return json.dumps(rows, indent=2)
        except Exception as e:
            raise Exception(f"CSV extraction failed: {str(e)}")
    
    def _extract_json(self, file_path: Path) -> str:
        """Extract data from JSON files."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            return json.dumps(data, indent=2)
        except Exception as e:
            raise Exception(f"JSON extraction failed: {str(e)}")
    
    def _extract_html(self, file_path: Path) -> str:
        """Extract text from HTML files."""
        try:
            from html.parser import HTMLParser
            
            class TextExtractor(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.text = []
                
                def handle_data(self, data):
                    text = data.strip()
                    if text:
                        self.text.append(text)
            
            with open(file_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
            
            parser = TextExtractor()
            parser.feed(html_content)
            return " ".join(parser.text)
        except Exception as e:
            raise Exception(f"HTML extraction failed: {str(e)}")
    
    def _extract_md(self, file_path: Path) -> str:
        """Extract text from Markdown files."""
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    
    def _get_file_metadata(self, file_path: Path) -> Dict[str, Any]:
        """Get metadata about the file."""
        stat = file_path.stat()
        return {
            "file_path": str(file_path),
            "file_size_kb": round(stat.st_size / 1024, 2),
            "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
            "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "extension": file_path.suffix
        }
    
    def extract_batch(self, file_paths: list) -> list:
        """
        Extract content from multiple files.
        
        Args:
            file_paths: List of file paths
            
        Returns:
            List of extraction results
        """
        results = []
        for file_path in file_paths:
            result = self.extract(file_path)
            results.append(result)
        return results
    
    def _generate_file_id(self, file_path: str) -> str:
        """Generate a unique ID for a file based on its path and modification time."""
        file_path_obj = Path(file_path)
        stat = file_path_obj.stat()
        id_string = f"{file_path}{stat.st_mtime}{stat.st_size}"
        return hashlib.md5(id_string.encode()).hexdigest()
    
    def save_extraction(self, extraction_result: Dict[str, Any], 
                       save_format: str = "all", 
                       save_to_db: bool = True) -> Dict[str, Any]:
        """
        Save extraction result to files and/or database.
        
        Args:
            extraction_result: Result from extract() method
            save_format: Format to save ('json', 'text', 'all')
            save_to_db: Whether to save to database
            
        Returns:
            Dictionary with save locations and status
        """
        if not extraction_result.get('success'):
            return {
                "success": False,
                "error": "Cannot save failed extraction",
                "extraction_status": extraction_result.get('error')
            }
        
        file_name = extraction_result.get('file_name', 'unknown')
        file_id = self._generate_file_id(extraction_result.get('file_path', file_name))
        
        save_paths = {}
        
        try:
            # Save as JSON
            if save_format in ['json', 'all']:
                json_path = self._save_as_json(extraction_result, file_id)
                save_paths['json'] = json_path
            
            # Save as text
            if save_format in ['text', 'all']:
                text_path = self._save_as_text(extraction_result, file_id)
                save_paths['text'] = text_path
            
            # Save to database
            if save_to_db:
                db_saved = self._save_to_database(extraction_result, file_id)
                save_paths['database'] = db_saved
            
            return {
                "success": True,
                "file_id": file_id,
                "file_name": file_name,
                "saved_paths": save_paths,
                "saved_at": datetime.now().isoformat()
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"Error saving extraction: {str(e)}",
                "file_id": file_id
            }
    
    def _save_as_json(self, extraction_result: Dict[str, Any], file_id: str) -> str:
        """Save extraction result as JSON file."""
        file_name = extraction_result.get('file_name', 'unknown').replace('.', '_')
        json_file = Path(f"{self.results_dir}/json/{file_name}_{file_id}.json")
        
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(extraction_result, f, indent=2, default=str)
        
        return str(json_file)
    
    def _save_as_text(self, extraction_result: Dict[str, Any], file_id: str) -> str:
        """Save extracted content as text file."""
        file_name = extraction_result.get('file_name', 'unknown').replace('.', '_')
        text_file = Path(f"{self.results_dir}/text/{file_name}_{file_id}.txt")
        
        content = extraction_result.get('content', 'No content extracted')
        
        with open(text_file, 'w', encoding='utf-8') as f:
            f.write(f"File: {extraction_result.get('file_name')}\n")
            f.write(f"Type: {extraction_result.get('file_type')}\n")
            f.write(f"Extracted: {extraction_result.get('extracted_at')}\n")
            f.write("=" * 80 + "\n\n")
            f.write(str(content))
        
        return str(text_file)
    
    def _save_to_database(self, extraction_result: Dict[str, Any], file_id: str) -> str:
        """Save extraction record to SQLite database."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            content = extraction_result.get('content', '')
            content_summary = str(content)[:500]  # Store first 500 chars as summary
            
            cursor.execute('''
                INSERT OR REPLACE INTO extractions 
                (file_id, file_name, file_type, file_path, file_size_bytes, 
                 content, content_summary, metadata, extracted_at, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                file_id,
                extraction_result.get('file_name'),
                extraction_result.get('file_type'),
                extraction_result.get('metadata', {}).get('file_path'),
                extraction_result.get('file_size_bytes'),
                content if isinstance(content, str) else json.dumps(content),
                content_summary,
                json.dumps(extraction_result.get('metadata', {})),
                extraction_result.get('extracted_at'),
                'success'
            ))
            
            conn.commit()
            conn.close()
            
            return f"Saved to database: {self.db_path}"
        except Exception as e:
            raise Exception(f"Database save failed: {str(e)}")
    
    def extract_and_save(self, file_path: str, save_format: str = "all") -> Dict[str, Any]:
        """
        Extract content from file and automatically save results.
        
        Args:
            file_path: Path to the file
            save_format: Format to save ('json', 'text', 'all')
            
        Returns:
            Combined extraction and save result
        """
        # Extract
        extraction_result = self.extract(file_path)
        
        if not extraction_result.get('success'):
            return extraction_result
        
        # Save
        save_result = self.save_extraction(extraction_result, save_format)
        
        return {
            **extraction_result,
            "saved": save_result
        }
    
    def get_saved_extractions(self, file_id: Optional[str] = None) -> list:
        """
        Retrieve saved extractions from database.
        
        Args:
            file_id: Optional specific file ID to retrieve
            
        Returns:
            List of extraction records
        """
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            if file_id:
                cursor.execute('SELECT * FROM extractions WHERE file_id = ?', (file_id,))
            else:
                cursor.execute('SELECT * FROM extractions ORDER BY extracted_at DESC')
            
            rows = cursor.fetchall()
            conn.close()
            
            results = []
            for row in rows:
                record = dict(row)
                # Parse JSON fields
                if record.get('metadata'):
                    try:
                        record['metadata'] = json.loads(record['metadata'])
                    except:
                        pass
                results.append(record)
            
            return results
        except Exception as e:
            return []
    
    def delete_extraction(self, file_id: str) -> Dict[str, Any]:
        """Delete an extraction record and associated files."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('DELETE FROM extractions WHERE file_id = ?', (file_id,))
            conn.commit()
            conn.close()
            
            # Delete associated JSON and text files
            json_files = list(Path(f"{self.results_dir}/json").glob(f"*{file_id}.json"))
            text_files = list(Path(f"{self.results_dir}/text").glob(f"*{file_id}.txt"))
            
            for file in json_files + text_files:
                file.unlink()
            
            return {
                "success": True,
                "file_id": file_id,
                "message": "Extraction deleted"
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get statistics about saved extractions."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT COUNT(*) FROM extractions')
            total_extractions = cursor.fetchone()[0]
            
            cursor.execute('SELECT file_type, COUNT(*) FROM extractions GROUP BY file_type')
            by_type = dict(cursor.fetchall())
            
            cursor.execute('SELECT SUM(file_size_bytes) FROM extractions')
            total_size = cursor.fetchone()[0] or 0
            
            conn.close()
            
            return {
                "total_extractions": total_extractions,
                "by_file_type": by_type,
                "total_size_bytes": total_size,
                "total_size_mb": round(total_size / (1024 * 1024), 2)
            }
        except Exception as e:
            return {"error": str(e)}


def main():
    """Example usage of the FileExtractor."""
    import sys
    
    # Create extractor instance
    extractor = FileExtractor()
    
    # Example: Extract from a specific file
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        result = extractor.extract(file_path)
        print(json.dumps(result, indent=2))
    else:
        print("Usage: python extract.py <file_path>")
        print("\nSupported formats:")
        for fmt, desc in FileExtractor.SUPPORTED_FORMATS.items():
            print(f"  {fmt:8} - {desc}")
        print("\nExample: python extract.py document.pdf")


if __name__ == "__main__":
    main()
